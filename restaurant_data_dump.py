from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

def validate_data(data):
    try:
        opening = datetime.strptime(data["openingTime"], "%H:%M")
        closing = datetime.strptime(data["closingTime"], "%H:%M")

        if closing <= opening:
            closing = closing.replace(day=opening.day + 1)

        if data["numBags"] <= 0:
            print("Error: Number of bags must be positive.")
            return False

        if data["pricePerBag"] <= 0:
            print("Error: Price per bag must be positive.")
            return False

        if data["actualNumBags"] < 0:
            print("Error: Actual number of bags cannot be negative.")
            return False

        return True

    except ValueError as e:
        print("Invalid input format:", e)
        return False


def update_restaurant_data(data, filename="restaurant_data.xlsx"):

    if not validate_data(data):
        print("Data is NOT valid. Nothing was saved.")
        return

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Restaurants"
        ws.append([
            "ID", "Opening Time", "Closing Time",
            "Num Bags", "Price per Bag", "Actual Num Bags"
        ])
        wb.save(filename)

    wb = load_workbook(filename)
    ws = wb.active
    ws.append([
        data["id"],
        data["openingTime"],
        data["closingTime"],
        data["numBags"],
        data["pricePerBag"],
        data["actualNumBags"]
    ])
    wb.save(filename)

    print("Restaurant data saved successfully!")

def get_restaurant_input():
    print("\n--- Enter Restaurant Data ---")

    data = {}

    data["id"] = input("Restaurant ID: ")

    opening = input("Opening Time (HH:MM AM/PM): ")
    closing = input("Closing Time (HH:MM AM/PM): ")

    data["openingTime"] = datetime.strptime(opening, "%I:%M %p").strftime("%H:%M")
    data["closingTime"] = datetime.strptime(closing, "%I:%M %p").strftime("%H:%M")

    data["numBags"] = int(input("Number of bags: "))
    data["pricePerBag"] = float(input("Price per bag: "))
    data["actualNumBags"] = int(input("Actual number of bags: "))

    return data

if __name__ == "__main__":
    restaurant_data = get_restaurant_input()
    update_restaurant_data(restaurant_data)
import sqlite3
from datetime import datetime
import json

def initialize_db():
    conn = sqlite3.connect("tgtg.db")
    cursor = conn.cursor()
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Restaurant (
            restaurant_id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            location TEXT NOT NULL,
            num_of_bags INTEGER NOT NULL,
            remaining_bags INTEGER NOT NULL,
            overall_rating REAL NOT NULL,
            opening_time TEXT NOT NULL,
            closing_time TEXT NOT NULL
        )
    """)
    
    conn.commit()
    return conn, cursor

def validate_time(time_str):
    try:
        return datetime.strptime(time_str, "%I:%M %p")
    except ValueError:
        return None

def validate_restaurant_data(data):
    errors = []

    if not data["name"]:
        errors.append("Restaurant name is required.")

    if not data["location"]:
        errors.append("At least one location is required.")

    if data["num_of_bags"] <= 0:
        errors.append("Number of bags must be > 0.")

    if data["remaining_bags"] < 0:
        errors.append("Remaining bags cannot be negative.")

    if data["overall_rating"] < 0 or data["overall_rating"] > 5:
        errors.append("Overall rating must be between 0 and 5.")

    opening = validate_time(data["opening_time"])
    closing = validate_time(data["closing_time"])

    if not opening:
        errors.append("Opening time invalid. Use HH:MM AM/PM format.")
    if not closing:
        errors.append("Closing time invalid. Use HH:MM AM/PM format.")

    if opening and closing:
        if closing <= opening:
            closing = closing.replace(day=closing.day + 1)
        if (closing - opening).total_seconds() <= 0:
            errors.append("Closing time must be after opening time.")

    return errors

def insert_restaurant(cursor, data):
    cursor.execute("""
        INSERT INTO Restaurant
        (restaurant_id, name, location, num_of_bags, remaining_bags,
        overall_rating, opening_time, closing_time)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data["restaurant_id"],
        data["name"],
        json.dumps(list(data["location"])),
        data["num_of_bags"],
        data["remaining_bags"],
        data["overall_rating"],
        data["opening_time"],
        data["closing_time"]
    ))

def main():
    conn, cursor = initialize_db()
    
    print("--- Enter Restaurant Data ---")
    data = {
        "restaurant_id": int(input("Restaurant ID: ")),
        "name": input("Restaurant Name: ").strip(),
        "location": set(l.strip() for l in input("Locations (comma-separated): ").split(",")),
        "num_of_bags": int(input("Total number of bags: ")),
        "remaining_bags": int(input("Remaining number of bags: ")),
        "overall_rating": float(input("Overall rating (0-5): ")),
        "opening_time": input("Opening Time (HH:MM AM/PM): ").strip(),
        "closing_time": input("Closing Time (HH:MM AM/PM): ").strip()
    }
    
    errors = validate_restaurant_data(data)
    if errors:
        print("Errors in input:")
        for err in errors:
            print(" -", err)
        print("Restaurant NOT saved.")
        return

    try:
        insert_restaurant(cursor, data)
        conn.commit()
        print("Restaurant successfully saved!")
    except sqlite3.IntegrityError:
        print("Error: Restaurant ID already exists.")
    
    conn.close()

if __name__ == "__main__":
    main()
